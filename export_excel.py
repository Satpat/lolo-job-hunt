#!/usr/bin/env python3
"""
Convert data/businesses.json into a CSV and an Excel workbook for Lolo to
track applications. The .xlsx keeps an AutoFilter on the header row, which
Google Sheets preserves as a native filter when the file is uploaded/converted.

Usage:
    python3 export_excel.py
"""

import csv
import json
from pathlib import Path
from urllib.parse import quote, urlparse

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent
DATA_JSON = BASE / "data" / "businesses.json"
CSV_OUT = BASE / "data" / "businesses.csv"
XLSX_OUT = BASE / "Bundoora_Business_Directory.xlsx"

GROUP_OF = {
    "Cafe": "Hospo",
    "Restaurant/Takeaway": "Hospo",
    "Fast Food": "Hospo",
    "Supermarket": "Grocery",
    "Retail": "Retail",
    "Retail (general)": "Retail",
    "Medical Clinic": "Health",
    "Hospital": "Health",
    "Dentist": "Health",
    "Allied Health": "Health",
    "Pharmacy": "Health",
    "Gym": "Fitness",
    "Salon/Beauty": "Fitness",
    "Childcare": "Family",
    "Cinema": "Family",
}

CATEGORY_DISPLAY = {
    "Retail (general)": "Retail",
    "Restaurant/Takeaway": "Restaurant",
    "Salon/Beauty": "Beauty",
}

COLUMNS = [
    "Business Name",
    "Group",
    "Category",
    "Suburb",
    "Distance (km)",
    "Address",
    "Phone",
    "Website",
    "Gmap link",
    "Applied?",
    "Notes",
]

SOCIAL_LABELS = {
    "instagram.com": "Instagram",
    "facebook.com": "Facebook",
    "fb.com": "Facebook",
    "fb.me": "Facebook",
    "tiktok.com": "TikTok",
    "twitter.com": "Twitter",
    "x.com": "Twitter",
    "linkedin.com": "LinkedIn",
    "youtube.com": "YouTube",
    "youtu.be": "YouTube",
    "pinterest.com": "Pinterest",
    "snapchat.com": "Snapchat",
    "wa.me": "WhatsApp",
    "whatsapp.com": "WhatsApp",
}


def website_label(url: str, business_name: str) -> str:
    """Business name for a regular site; platform name for a social link."""
    if not url:
        return ""
    host = urlparse(url).netloc.lower()
    if host.startswith("www."):
        host = host[4:]
    for domain, label in SOCIAL_LABELS.items():
        if host == domain or host.endswith("." + domain):
            return label
    return business_name


def gmap_url(place_id: str, business_name: str) -> str:
    """Link to the business's own Google Maps listing, not raw GPS coords."""
    if not place_id:
        return ""
    return f"https://www.google.com/maps/search/?api=1&query={quote(business_name)}&query_place_id={place_id}"


def load_data():
    if not DATA_JSON.exists():
        raise SystemExit(f"Missing {DATA_JSON} - run fetch_businesses.py first.")
    return json.loads(DATA_JSON.read_text())


def to_row(b: dict) -> list:
    cat = b.get("category") or ""
    name = b.get("name") or ""
    return [
        name,
        GROUP_OF.get(cat, ""),
        CATEGORY_DISPLAY.get(cat, cat),
        b.get("suburb") or "",
        b.get("distance_km"),
        b.get("address") or "",
        b.get("phone") or "",
        b.get("website") or "",
        gmap_url(b.get("place_id"), name),
        "",  # Applied?
        "",  # Notes
    ]


def write_csv(rows):
    with CSV_OUT.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(COLUMNS)
        writer.writerows(rows)
    print(f"Wrote {CSV_OUT}")


HYPERLINK_FONT = Font(color="0563C1", underline="single")


def write_xlsx(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bundoora Directory"

    ws.append(COLUMNS)
    header_fill = PatternFill(start_color="8E2A52", end_color="8E2A52", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font

    website_col = COLUMNS.index("Website") + 1
    gmap_col = COLUMNS.index("Gmap link") + 1

    for row in rows:
        ws.append(row)
        r = ws.max_row
        name = row[0]

        website_val = row[website_col - 1]
        if website_val:
            cell = ws.cell(row=r, column=website_col)
            cell.value = website_label(website_val, name)
            cell.hyperlink = website_val
            cell.font = HYPERLINK_FONT

        gmap_val = row[gmap_col - 1]
        if gmap_val:
            cell = ws.cell(row=r, column=gmap_col)
            cell.value = "Directions"
            cell.hyperlink = gmap_val
            cell.font = HYPERLINK_FONT

    widths = [28, 20, 18, 16, 12, 40, 16, 22, 12, 10, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"

    wb.save(XLSX_OUT)
    print(f"Wrote {XLSX_OUT}")


def main():
    data = load_data()
    data.sort(key=lambda b: (b.get("distance_km") is None, b.get("distance_km"), b.get("name") or ""))
    rows = [to_row(b) for b in data]
    write_csv(rows)
    write_xlsx(rows)
    print(f"Total businesses: {len(rows)}")


if __name__ == "__main__":
    main()
